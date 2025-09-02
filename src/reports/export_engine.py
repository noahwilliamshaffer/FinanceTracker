"""
Data Export and Report Generation Engine
Professional Excel, PDF, and CSV export capabilities with templating and scheduling
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union
import io
import base64
import json
import logging
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
import tempfile
import zipfile

# Excel and PDF libraries
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference, BarChart
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not available - Excel export disabled")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("reportlab not available - PDF export disabled")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExportFormat(Enum):
    """Supported export formats"""
    EXCEL = "excel"
    PDF = "pdf"
    CSV = "csv"
    JSON = "json"

class ReportType(Enum):
    """Types of reports"""
    PORTFOLIO_SUMMARY = "portfolio_summary"
    RISK_REPORT = "risk_report"
    PNL_ATTRIBUTION = "pnl_attribution"
    MARKET_DATA = "market_data"
    TRADE_BLOTTER = "trade_blotter"
    COMPLIANCE_REPORT = "compliance_report"
    EXECUTIVE_SUMMARY = "executive_summary"

@dataclass
class ExportRequest:
    """Export request configuration"""
    report_type: ReportType
    format: ExportFormat
    title: str
    subtitle: str = ""
    data: Dict[str, Any] = field(default_factory=dict)
    parameters: Dict[str, Any] = field(default_factory=dict)
    template_name: Optional[str] = None
    include_charts: bool = True
    user_id: str = ""
    timestamp: datetime = field(default_factory=datetime.now)

@dataclass
class ExportResult:
    """Export operation result"""
    success: bool
    file_path: Optional[str] = None
    file_data: Optional[bytes] = None
    file_name: str = ""
    content_type: str = ""
    size_bytes: int = 0
    error_message: str = ""
    generation_time_seconds: float = 0.0

class ExcelExporter:
    """Professional Excel export with formatting and charts"""
    
    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        # Define styles
        self.header_font = Font(bold=True, size=14, color="FFFFFF")
        self.subheader_font = Font(bold=True, size=12)
        self.data_font = Font(size=10)
        self.number_font = Font(size=10)
        
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.right_alignment = Alignment(horizontal="right", vertical="center")
        
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def export_portfolio_summary(self, request: ExportRequest) -> ExportResult:
        """Export portfolio summary to Excel"""
        try:
            start_time = datetime.now()
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            self._create_summary_sheet(wb, request.data)
            
            # Create positions sheet
            if 'positions' in request.data:
                self._create_positions_sheet(wb, request.data['positions'])
            
            # Create risk metrics sheet
            if 'risk_metrics' in request.data:
                self._create_risk_sheet(wb, request.data['risk_metrics'])
            
            # Create P&L attribution sheet
            if 'pnl_attribution' in request.data:
                self._create_pnl_sheet(wb, request.data['pnl_attribution'])
            
            # Save to bytes
            file_data = io.BytesIO()
            wb.save(file_data)
            file_data.seek(0)
            
            generation_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=True,
                file_data=file_data.getvalue(),
                file_name=f"{request.title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                size_bytes=len(file_data.getvalue()),
                generation_time_seconds=generation_time
            )
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return ExportResult(
                success=False,
                error_message=str(e)
            )
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create portfolio summary sheet"""
        ws = wb.create_sheet("Portfolio Summary")
        
        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "Portfolio Summary Report"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = self.center_alignment
        
        # Date
        ws.merge_cells("A2:F2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].alignment = self.center_alignment
        
        # Key metrics
        row = 4
        summary_data = data.get('summary', {})
        
        metrics = [
            ("Total Market Value", summary_data.get('total_market_value', 0), "${:,.2f}"),
            ("Total P&L", summary_data.get('total_pnl', 0), "${:,.2f}"),
            ("Unrealized P&L", summary_data.get('total_unrealized_pnl', 0), "${:,.2f}"),
            ("Realized P&L", summary_data.get('total_realized_pnl', 0), "${:,.2f}"),
            ("Portfolio Duration", summary_data.get('portfolio_duration', 0), "{:.2f}"),
            ("Portfolio Yield", summary_data.get('portfolio_yield', 0), "{:.2%}"),
            ("95% VaR", summary_data.get('var_95', 0), "${:,.2f}"),
            ("Number of Positions", summary_data.get('positions_count', 0), "{:,}")
        ]
        
        for i, (label, value, fmt) in enumerate(metrics):
            ws[f"A{row + i}"] = label
            ws[f"A{row + i}"].font = self.subheader_font
            
            ws[f"B{row + i}"] = fmt.format(value)
            ws[f"B{row + i}"].font = self.number_font
            ws[f"B{row + i}"].alignment = self.right_alignment
            
            # Add color coding for P&L
            if "P&L" in label and value != 0:
                fill_color = "C6EFCE" if value > 0 else "FFC7CE"  # Green for positive, red for negative
                ws[f"B{row + i}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_positions_sheet(self, wb: openpyxl.Workbook, positions_data: List[Dict]):
        """Create positions detail sheet"""
        ws = wb.create_sheet("Positions")
        
        if not positions_data:
            ws["A1"] = "No position data available"
            return
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(positions_data)
        
        # Headers
        headers = [
            "CUSIP", "Security Name", "Quantity (MM)", "Avg Price", "Current Price",
            "Market Value", "Unrealized P&L", "Duration", "Yield", "Asset Class"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            data_row = [
                row.get('cusip', ''),
                row.get('security_name', ''),
                row.get('quantity', 0),
                row.get('average_price', 0),
                row.get('current_price', 0),
                row.get('market_value', 0),
                row.get('unrealized_pnl', 0),
                row.get('duration', 0),
                row.get('yield_rate', 0),
                row.get('asset_class', '')
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border
                
                # Format numbers
                if col in [3, 4, 5, 6, 7, 8, 9]:  # Numeric columns
                    if col in [6, 7]:  # Dollar amounts
                        cell.number_format = '"$"#,##0.00'
                    elif col == 9:  # Percentage
                        cell.number_format = '0.00%'
                    else:  # Other numbers
                        cell.number_format = '#,##0.00'
                
                # Alternate row coloring
                if row_idx % 2 == 0:
                    cell.fill = self.alternate_fill
                
                # P&L color coding
                if col == 7 and isinstance(value, (int, float)):  # Unrealized P&L
                    if value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_risk_sheet(self, wb: openpyxl.Workbook, risk_data: Dict[str, Any]):
        """Create risk metrics sheet"""
        ws = wb.create_sheet("Risk Metrics")
        
        # Title
        ws["A1"] = "Risk Analysis"
        ws["A1"].font = Font(bold=True, size=14)
        
        row = 3
        
        # VaR Analysis
        if 'var_analysis' in risk_data:
            ws[f"A{row}"] = "Value at Risk Analysis"
            ws[f"A{row}"].font = self.subheader_font
            row += 1
            
            var_data = risk_data['var_analysis']
            var_metrics = [
                ("95% VaR", var_data.get('var_95', 0)),
                ("99% VaR", var_data.get('var_99', 0)),
                ("Expected Shortfall 95%", var_data.get('es_95', 0)),
                ("Expected Shortfall 99%", var_data.get('es_99', 0))
            ]
            
            for label, value in var_metrics:
                ws[f"A{row}"] = label
                ws[f"B{row}"] = f"${value:,.2f}"
                ws[f"B{row}"].number_format = '"$"#,##0.00'
                row += 1
            
            row += 1
        
        # Duration Analysis
        if 'duration_analysis' in risk_data:
            ws[f"A{row}"] = "Duration Analysis"
            ws[f"A{row}"].font = self.subheader_font
            row += 1
            
            duration_data = risk_data['duration_analysis']
            duration_metrics = [
                ("Portfolio Duration", duration_data.get('portfolio_duration', 0), "{:.2f}"),
                ("Portfolio Convexity", duration_data.get('portfolio_convexity', 0), "{:.2f}"),
                ("DV01", duration_data.get('dv01', 0), "${:,.2f}")
            ]
            
            for label, value, fmt in duration_metrics:
                ws[f"A{row}"] = label
                ws[f"B{row}"] = fmt.format(value)
                row += 1
    
    def _create_pnl_sheet(self, wb: openpyxl.Workbook, pnl_data: Dict[str, Any]):
        """Create P&L attribution sheet"""
        ws = wb.create_sheet("P&L Attribution")
        
        # Title
        ws["A1"] = "P&L Attribution Analysis"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Daily attribution
        if 'daily_attribution' in pnl_data:
            row = 3
            ws[f"A{row}"] = "Daily P&L Attribution"
            ws[f"A{row}"].font = self.subheader_font
            row += 1
            
            attribution = pnl_data['daily_attribution']
            factors = [
                ("Carry P&L", attribution.get('carry_pnl', 0)),
                ("Price P&L", attribution.get('price_pnl', 0)),
                ("Curve P&L", attribution.get('curve_pnl', 0)),
                ("Spread P&L", attribution.get('spread_pnl', 0)),
                ("Other P&L", attribution.get('other_pnl', 0))
            ]
            
            # Headers
            ws[f"A{row}"] = "Factor"
            ws[f"B{row}"] = "P&L ($)"
            ws[f"C{row}"] = "Contribution %"
            
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            row += 1
            
            total_pnl = sum(value for _, value in factors)
            
            for factor, value in factors:
                ws[f"A{row}"] = factor
                ws[f"B{row}"] = value
                ws[f"B{row}"].number_format = '"$"#,##0.00'
                
                if total_pnl != 0:
                    contribution = (value / total_pnl) * 100
                    ws[f"C{row}"] = contribution
                    ws[f"C{row}"].number_format = '0.0%'
                
                # Color coding
                if value > 0:
                    ws[f"B{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value < 0:
                    ws[f"B{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row += 1

class PDFExporter:
    """Professional PDF report generation"""
    
    def __init__(self):
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")
        
        self.styles = getSampleStyleSheet()
        
        # Custom styles
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Title'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading1'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#366092')
        )
    
    def export_executive_summary(self, request: ExportRequest) -> ExportResult:
        """Export executive summary to PDF"""
        try:
            start_time = datetime.now()
            
            # Create PDF buffer
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            
            # Build content
            story = []
            
            # Title
            title = Paragraph(request.title, self.title_style)
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Subtitle with date
            subtitle_text = f"{request.subtitle}<br/><font size='10'>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font>"
            subtitle = Paragraph(subtitle_text, self.styles['Normal'])
            story.append(subtitle)
            story.append(Spacer(1, 20))
            
            # Executive Summary
            self._add_executive_summary_section(story, request.data)
            
            # Key Metrics
            self._add_key_metrics_section(story, request.data)
            
            # Risk Summary
            self._add_risk_summary_section(story, request.data)
            
            # Build PDF
            doc.build(story)
            
            # Get PDF data
            buffer.seek(0)
            file_data = buffer.getvalue()
            buffer.close()
            
            generation_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=True,
                file_data=file_data,
                file_name=f"{request.title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                content_type="application/pdf",
                size_bytes=len(file_data),
                generation_time_seconds=generation_time
            )
            
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            return ExportResult(
                success=False,
                error_message=str(e)
            )
    
    def _add_executive_summary_section(self, story: List, data: Dict[str, Any]):
        """Add executive summary section"""
        story.append(Paragraph("Executive Summary", self.heading_style))
        
        summary_data = data.get('summary', {})
        
        # Key highlights
        highlights = [
            f"Portfolio market value: ${summary_data.get('total_market_value', 0):,.2f}",
            f"Total P&L: ${summary_data.get('total_pnl', 0):,.2f}",
            f"Portfolio duration: {summary_data.get('portfolio_duration', 0):.2f} years",
            f"95% VaR: ${summary_data.get('var_95', 0):,.2f}",
            f"Active positions: {summary_data.get('positions_count', 0)}"
        ]
        
        for highlight in highlights:
            story.append(Paragraph(f"• {highlight}", self.styles['Normal']))
        
        story.append(Spacer(1, 20))
    
    def _add_key_metrics_section(self, story: List, data: Dict[str, Any]):
        """Add key metrics table"""
        story.append(Paragraph("Key Portfolio Metrics", self.heading_style))
        
        summary_data = data.get('summary', {})
        
        # Create metrics table
        table_data = [
            ['Metric', 'Value'],
            ['Total Market Value', f"${summary_data.get('total_market_value', 0):,.2f}"],
            ['Unrealized P&L', f"${summary_data.get('total_unrealized_pnl', 0):,.2f}"],
            ['Realized P&L', f"${summary_data.get('total_realized_pnl', 0):,.2f}"],
            ['Portfolio Duration', f"{summary_data.get('portfolio_duration', 0):.2f}"],
            ['Portfolio Yield', f"{summary_data.get('portfolio_yield', 0):.2%}"],
            ['95% VaR', f"${summary_data.get('var_95', 0):,.2f}"],
            ['99% VaR', f"${summary_data.get('var_99', 0):,.2f}"]
        ]
        
        table = Table(table_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
    
    def _add_risk_summary_section(self, story: List, data: Dict[str, Any]):
        """Add risk summary section"""
        story.append(Paragraph("Risk Summary", self.heading_style))
        
        risk_text = """
        The portfolio risk metrics indicate the current exposure to market movements.
        Value at Risk (VaR) represents the potential loss over a one-day period at the specified confidence level.
        Duration measures the portfolio's sensitivity to interest rate changes.
        """
        
        story.append(Paragraph(risk_text, self.styles['Normal']))
        story.append(Spacer(1, 12))

class CSVExporter:
    """Simple CSV export functionality"""
    
    @staticmethod
    def export_dataframe(df: pd.DataFrame, filename: str) -> ExportResult:
        """Export DataFrame to CSV"""
        try:
            start_time = datetime.now()
            
            # Convert to CSV
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False)
            csv_data = csv_buffer.getvalue().encode('utf-8')
            
            generation_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=True,
                file_data=csv_data,
                file_name=filename if filename.endswith('.csv') else f"{filename}.csv",
                content_type="text/csv",
                size_bytes=len(csv_data),
                generation_time_seconds=generation_time
            )
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return ExportResult(
                success=False,
                error_message=str(e)
            )

class ReportEngine:
    """Main report generation engine"""
    
    def __init__(self):
        self.excel_exporter = ExcelExporter() if EXCEL_AVAILABLE else None
        self.pdf_exporter = PDFExporter() if PDF_AVAILABLE else None
        self.csv_exporter = CSVExporter()
        
        # Report templates
        self.templates = {
            ReportType.PORTFOLIO_SUMMARY: self._generate_portfolio_summary,
            ReportType.RISK_REPORT: self._generate_risk_report,
            ReportType.EXECUTIVE_SUMMARY: self._generate_executive_summary,
            ReportType.TRADE_BLOTTER: self._generate_trade_blotter
        }
    
    def generate_report(self, request: ExportRequest) -> ExportResult:
        """Generate report based on request"""
        logger.info(f"Generating {request.report_type.value} report in {request.format.value} format")
        
        try:
            # Get report generator
            generator = self.templates.get(request.report_type)
            if not generator:
                return ExportResult(
                    success=False,
                    error_message=f"Unsupported report type: {request.report_type.value}"
                )
            
            # Generate report data
            report_data = generator(request.data, request.parameters)
            request.data = report_data
            
            # Export based on format
            if request.format == ExportFormat.EXCEL:
                if not self.excel_exporter:
                    return ExportResult(success=False, error_message="Excel export not available")
                return self.excel_exporter.export_portfolio_summary(request)
            
            elif request.format == ExportFormat.PDF:
                if not self.pdf_exporter:
                    return ExportResult(success=False, error_message="PDF export not available")
                return self.pdf_exporter.export_executive_summary(request)
            
            elif request.format == ExportFormat.CSV:
                # Convert main data to DataFrame
                main_data = report_data.get('positions', [])
                if main_data:
                    df = pd.DataFrame(main_data)
                    return self.csv_exporter.export_dataframe(df, request.title)
                else:
                    return ExportResult(success=False, error_message="No data available for CSV export")
            
            elif request.format == ExportFormat.JSON:
                json_data = json.dumps(report_data, default=str, indent=2).encode('utf-8')
                return ExportResult(
                    success=True,
                    file_data=json_data,
                    file_name=f"{request.title.replace(' ', '_')}.json",
                    content_type="application/json",
                    size_bytes=len(json_data)
                )
            
            else:
                return ExportResult(
                    success=False,
                    error_message=f"Unsupported export format: {request.format.value}"
                )
                
        except Exception as e:
            logger.error(f"Report generation failed: {e}")
            return ExportResult(
                success=False,
                error_message=str(e)
            )
    
    def _generate_portfolio_summary(self, data: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        """Generate portfolio summary report data"""
        return {
            'summary': data.get('portfolio_summary', {}),
            'positions': data.get('positions', []),
            'risk_metrics': data.get('risk_metrics', {}),
            'pnl_attribution': data.get('pnl_attribution', {}),
            'breakdown': data.get('position_breakdown', {})
        }
    
    def _generate_risk_report(self, data: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        """Generate risk report data"""
        return {
            'var_analysis': data.get('var_analysis', {}),
            'duration_analysis': data.get('duration_analysis', {}),
            'stress_testing': data.get('stress_testing', {}),
            'concentration_risk': data.get('concentration_risk', {})
        }
    
    def _generate_executive_summary(self, data: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        """Generate executive summary data"""
        return {
            'summary': data.get('portfolio_summary', {}),
            'key_highlights': data.get('key_highlights', []),
            'risk_summary': data.get('risk_summary', {}),
            'performance_metrics': data.get('performance_metrics', {})
        }
    
    def _generate_trade_blotter(self, data: Dict[str, Any], params: Dict[str, Any]) -> Dict[str, Any]:
        """Generate trade blotter data"""
        return {
            'trades': data.get('trades', []),
            'summary_stats': data.get('trade_summary', {}),
            'counterparty_breakdown': data.get('counterparty_breakdown', {})
        }
    
    def get_available_formats(self) -> List[str]:
        """Get list of available export formats"""
        formats = [ExportFormat.JSON.value, ExportFormat.CSV.value]
        
        if self.excel_exporter:
            formats.append(ExportFormat.EXCEL.value)
        
        if self.pdf_exporter:
            formats.append(ExportFormat.PDF.value)
        
        return formats
    
    def create_bulk_export(self, requests: List[ExportRequest]) -> ExportResult:
        """Create bulk export as ZIP file"""
        try:
            start_time = datetime.now()
            
            # Create ZIP buffer
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for request in requests:
                    result = self.generate_report(request)
                    
                    if result.success and result.file_data:
                        zip_file.writestr(result.file_name, result.file_data)
                    else:
                        # Add error report
                        error_content = f"Export failed: {result.error_message}"
                        zip_file.writestr(f"ERROR_{request.title}.txt", error_content)
            
            zip_data = zip_buffer.getvalue()
            generation_time = (datetime.now() - start_time).total_seconds()
            
            return ExportResult(
                success=True,
                file_data=zip_data,
                file_name=f"bulk_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                content_type="application/zip",
                size_bytes=len(zip_data),
                generation_time_seconds=generation_time
            )
            
        except Exception as e:
            logger.error(f"Bulk export failed: {e}")
            return ExportResult(
                success=False,
                error_message=str(e)
            )

# Example usage and testing
if __name__ == "__main__":
    # Create sample data
    sample_data = {
        'portfolio_summary': {
            'total_market_value': 150000000,
            'total_pnl': 2500000,
            'total_unrealized_pnl': 1800000,
            'total_realized_pnl': 700000,
            'portfolio_duration': 6.5,
            'portfolio_yield': 0.045,
            'var_95': 850000,
            'var_99': 1200000,
            'positions_count': 25
        },
        'positions': [
            {
                'cusip': '912828XG8',
                'security_name': '10Y Treasury Note',
                'quantity': 10.0,
                'average_price': 99.5,
                'current_price': 100.1,
                'market_value': 10010000,
                'unrealized_pnl': 60000,
                'duration': 8.5,
                'yield_rate': 0.045,
                'asset_class': 'treasury'
            },
            {
                'cusip': '912828YK0',
                'security_name': '2Y Treasury Note',
                'quantity': 5.0,
                'average_price': 101.2,
                'current_price': 100.8,
                'market_value': 5040000,
                'unrealized_pnl': -20000,
                'duration': 1.9,
                'yield_rate': 0.048,
                'asset_class': 'treasury'
            }
        ]
    }
    
    # Create report engine
    engine = ReportEngine()
    
    print(f"Available formats: {engine.get_available_formats()}")
    
    # Test CSV export
    csv_request = ExportRequest(
        report_type=ReportType.PORTFOLIO_SUMMARY,
        format=ExportFormat.CSV,
        title="Portfolio Positions",
        data=sample_data
    )
    
    csv_result = engine.generate_report(csv_request)
    if csv_result.success:
        print(f"✅ CSV export successful: {csv_result.file_name} ({csv_result.size_bytes} bytes)")
    else:
        print(f"❌ CSV export failed: {csv_result.error_message}")
    
    # Test JSON export
    json_request = ExportRequest(
        report_type=ReportType.PORTFOLIO_SUMMARY,
        format=ExportFormat.JSON,
        title="Portfolio Summary",
        data=sample_data
    )
    
    json_result = engine.generate_report(json_request)
    if json_result.success:
        print(f"✅ JSON export successful: {json_result.file_name} ({json_result.size_bytes} bytes)")
    else:
        print(f"❌ JSON export failed: {json_result.error_message}")
    
    # Test Excel export (if available)
    if EXCEL_AVAILABLE:
        excel_request = ExportRequest(
            report_type=ReportType.PORTFOLIO_SUMMARY,
            format=ExportFormat.EXCEL,
            title="Portfolio Analysis Report",
            subtitle="Comprehensive portfolio overview with risk metrics",
            data=sample_data
        )
        
        excel_result = engine.generate_report(excel_request)
        if excel_result.success:
            print(f"✅ Excel export successful: {excel_result.file_name} ({excel_result.size_bytes} bytes)")
            print(f"   Generation time: {excel_result.generation_time_seconds:.2f} seconds")
        else:
            print(f"❌ Excel export failed: {excel_result.error_message}")
    
    print("\n📊 Export engine ready!")
